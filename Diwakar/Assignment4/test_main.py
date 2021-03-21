import pytest
from selenium import webdriver
from assignment.loginpage import LoginPage
from assignment.dymanictablepage import DymanictablePage
import openpyxl


@pytest.fixture(autouse=True)
def setup():
    driver = webdriver.Chrome(executable_path='/home/diwakar/Desktop/chromedriver_linux64/chromedriver')
    yield driver
    driver.quit()


def test_login(setup):
    driver = setup
    driver.get("https://www.phptravels.net/login")
    driver.maximize_window()
    loginpage = LoginPage(driver)
    wbj_obj = openpyxl.load_workbook('./data.xlsx')
    sheet_obj = wbj_obj.active
    loginpage.enter_email(sheet_obj.cell(row=2, column=1).value)
    loginpage.enter_password(sheet_obj.cell(row=2, column=2).value)
    loginpage.click_login()


def test_dynamic_table(setup):
    driver = setup
    driver.get("http://demo.guru99.com/test/web-table-element.php")
    driver.maximize_window()
    dynamictablepage = DymanictablePage(driver)
    table = dynamictablepage.table
    header_list = table.find_elements_by_tag_name('th')
    print("")
    for header in header_list:
        text = header.text
        print(text, end='            ')

    rows_list = table.find_elements_by_tag_name('tr')
    for rows in rows_list:
        cols = rows.find_elements_by_tag_name('td')
        for col in cols:
            print(col.text, end='            ')
        print("")
